#!/usr/bin/env python3
"""Erzeugt die XLSX-Aktenstuecke fuer die Schaefer-Akte."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "xlsx"
OUT.mkdir(exist_ok=True)

THIN = Side(border_style="thin", color="D4D1CA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TEAL = PatternFill(start_color="01696F", end_color="01696F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
SUM_FILL = PatternFill(start_color="F7F6F2", end_color="F7F6F2", fill_type="solid")


def style_header(cell):
    cell.fill = TEAL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_cell(cell, bold=False, sum_=False):
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.border = BORDER
    if sum_:
        cell.fill = SUM_FILL


def auto_width(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_versorgungsausgleich():
    wb = Workbook()
    ws = wb.active
    ws.title = "Versorgungsausgleich"

    ws["A1"] = "Versorgungsausgleichsberechnung Schaefer ./. Schaefer"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="01696F")
    ws.merge_cells("A1:F1")

    ws["A2"] = "Az. AG Nuernberg-Fuerth 110 F 2418/21 ; Ehezeit 01.09.1987 - 30.04.2021"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="7A7974")
    ws.merge_cells("A2:F2")

    headers = ["Versorgungstraeger / Vertrag", "Anrechtinhaber", "Anrecht (ehezeitlich)", "Korresp. Kapital EUR", "Ausgleichswert", "Ausgleichsform"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    rows = [
        ("Bayerische Aerzteversorgung (Mitgl-Nr. 14-072 184/1958)", "Dr. M. Schaefer", "4.127,42 EUR/Monat", 412853.20, "2.063,71 EUR/Monat", "interne Teilung § 10"),
        ("DRV Nordbayern (Vers-Nr. 65 030258 S 010)", "Dr. M. Schaefer", "2,1834 EP", 16842.11, "1,0917 EP", "interne Teilung § 10"),
        ("Allianz Riester (Vertr. 0817 4118 22)", "Dr. M. Schaefer", "DK 18.473,42 EUR", 18473.42, "8.736,71 EUR (Kapital)", "externe Teilung § 14"),
        ("Wuerttembergische Direktvers (4471 9982 73)", "Dr. M. Schaefer", "DK 44.831,77 EUR", 44831.77, "22.165,89 EUR (Kapital)", "externe Teilung § 14"),
        ("BayLfF Beamtenversorgung Bayern", "Frau L. Schaefer", "1.850,00 EUR/Monat", 185000.00, "925,00 EUR/Monat", "interne Teilung § 10"),
        ("DRV Bund", "Frau L. Schaefer", "5,2000 EP (ehezeitlich; 12,5000 EP gesamt)", 40118.00, "2,6000 EP", "interne Teilung § 10"),
    ]
    for i, r in enumerate(rows, start=5):
        for j, val in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=val)
            style_cell(c, bold=(j == 1))
            if j == 4:
                c.number_format = "#,##0.00 [$EUR]"

    # Summen
    last = len(rows) + 4
    ws.cell(row=last + 2, column=1, value="Saldenuebersicht (informativ)").font = Font(name="Calibri", size=11, bold=True, color="01696F")
    ws.cell(row=last + 4, column=1, value="Ausgleichssumme (Kapital), Anteile Dr. M. Schaefer:")
    ws.cell(row=last + 4, column=4, value=412853.20/2 + 16842.11/2 + 8736.71 + 22165.89).number_format = "#,##0.00 [$EUR]"
    ws.cell(row=last + 5, column=1, value="Ausgleichssumme (Kapital), Anteile Frau L. Schaefer:")
    ws.cell(row=last + 5, column=4, value=185000.00/2 + 40118.00/2).number_format = "#,##0.00 [$EUR]"
    ws.cell(row=last + 7, column=1, value="Bei Ausschluss nach § 27 VersAusglG").font = Font(name="Calibri", size=10, bold=True, color="01696F")
    ws.cell(row=last + 8, column=1, value="Beide Anrechte bleiben in voller Hoehe beim urspruenglichen Inhaber.")

    auto_width(ws, [42, 22, 26, 22, 26, 26])

    # 2. Sheet: Halbteilungssimulation
    ws2 = wb.create_sheet("Halbteilung_Simulation")
    ws2["A1"] = "Halbteilungssimulation - monatliche Bruttorenten"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="01696F")
    ws2.merge_cells("A1:E1")

    sim_headers = ["Anrecht", "Inhaber vor VA EUR/M", "Empfaenger nach VA EUR/M", "Saldo Dr. M.", "Saldo Frau L."]
    for i, h in enumerate(sim_headers, start=1):
        style_header(ws2.cell(row=3, column=i, value=h))

    sim_rows = [
        ("Bayerische Aerzteversorgung", 4127.42, 2063.71, -2063.71, 2063.71),
        ("DRV (umgerechnet, ca.)", 75.65, 37.82, -37.82, 37.82),
        ("BayLfF Beamtenversorgung", 1850.00, 925.00, 925.00, -925.00),
        ("DRV Bund (ehezeitlich, ca.)", 180.02, 90.01, 90.01, -90.01),
    ]
    for i, r in enumerate(sim_rows, start=4):
        for j, v in enumerate(r, start=1):
            c = ws2.cell(row=i, column=j, value=v)
            style_cell(c, bold=(j == 1))
            if j > 1:
                c.number_format = "#,##0.00 [$EUR]"

    bottom = 4 + len(sim_rows)
    ws2.cell(row=bottom + 1, column=1, value="Netto-Saldo (Saldenubertragung)").font = BOLD_FONT
    ws2.cell(row=bottom + 1, column=4, value=sum(r[3] for r in sim_rows)).number_format = "#,##0.00 [$EUR]"
    ws2.cell(row=bottom + 1, column=5, value=sum(r[4] for r in sim_rows)).number_format = "#,##0.00 [$EUR]"

    ws2.cell(row=bottom + 3, column=1, value="Lesehilfe").font = Font(name="Calibri", size=10, bold=True, color="01696F")
    ws2.cell(row=bottom + 4, column=1, value="Negative Werte bedeuten Reduktion bei Dr. M. Schaefer.")
    ws2.cell(row=bottom + 5, column=1, value="Positive Werte bei Frau L. Schaefer bedeuten Zugewinn an Anwartschaft.")
    ws2.cell(row=bottom + 7, column=1, value="Wenn § 27 VersAusglG angewendet wird, entfaellt diese Tabelle (kein VA).").font = Font(name="Calibri", size=10, italic=True, color="7A7974")

    auto_width(ws2, [32, 22, 26, 18, 18])

    wb.save(OUT / "versorgungsausgleich_berechnung_schaefer.xlsx")


def build_liquiditaetsbetrachtung():
    wb = Workbook()
    ws = wb.active
    ws.title = "Immobilien"

    ws["A1"] = "Liquiditaetsbetrachtung Immobilien Frau L. Schaefer (Erbschaft Winterstein 2019)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="01696F")
    ws.merge_cells("A1:G1")

    ws["A2"] = "Stichtag 31.12.2021; Werte aus SV-Gutachten Weisbacher 18.10.2022 sowie Hausverwaltung Pirsch und Maerz"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="7A7974")
    ws.merge_cells("A2:G2")

    headers = ["Objekt", "Verkehrswert EUR", "Miet-Brutto p.M. EUR", "Bewirtschaftungskosten p.M. EUR", "Netto-Cashflow p.M. EUR", "Verwertbarkeit kurzfr.", "Beleihungsrahmen EUR"]
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=4, column=i, value=h))

    rows = [
        ("Mehrfamilienhaus Bamberg Promenadestr. 14 (12 WE)", 2795000, 12000, 4800, 7200, "kurzfr. 8-12% Abschlag", 1677000),
        ("Gewerbeimmobilie Forchheim Bayreuther Str. 88", 1510000, 8500, 3000, 5500, "mittelfr. 15-20% Abschlag", 906000),
        ("ETW Muenchen Hohenzollernstr. 47", 945000, 2800, 750, 2050, "kurzfr. 5-8% Abschlag", 567000),
    ]
    for i, r in enumerate(rows, start=5):
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=v)
            style_cell(c, bold=(j == 1))
            if j in (2, 3, 4, 5, 7):
                c.number_format = "#,##0 [$EUR]"

    last = 4 + len(rows)
    sum_row = last + 1
    ws.cell(row=sum_row, column=1, value="Summen").font = BOLD_FONT
    for j in (2, 3, 4, 5, 7):
        c = ws.cell(row=sum_row, column=j, value=sum(r[j-1] for r in rows))
        style_cell(c, bold=True, sum_=True)
        c.number_format = "#,##0 [$EUR]"

    ws.cell(row=sum_row + 2, column=1, value="Zusatzposten Vermoegen").font = Font(name="Calibri", size=11, bold=True, color="01696F")
    ws.cell(row=sum_row + 3, column=1, value="Wertpapierdepot DZ Privatbank (4119 002 71)")
    ws.cell(row=sum_row + 3, column=2, value=1200000).number_format = "#,##0 [$EUR]"
    ws.cell(row=sum_row + 4, column=1, value="Festgeldkonten Sparkasse Bamberg")
    ws.cell(row=sum_row + 4, column=2, value=450000).number_format = "#,##0 [$EUR]"

    ws.cell(row=sum_row + 6, column=1, value="Gesamt-Vermoegen").font = Font(name="Calibri", size=11, bold=True, color="01696F")
    ws.cell(row=sum_row + 6, column=2, value=sum(r[1] for r in rows) + 1200000 + 450000).number_format = "#,##0 [$EUR]"

    auto_width(ws, [44, 18, 18, 24, 22, 26, 22])

    # 2. Sheet: Beleihbarkeit / Substanzrisiken
    ws2 = wb.create_sheet("Beleihung_Substanz")
    ws2["A1"] = "Beleihbarkeit und Substanzrisiken"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="01696F")
    ws2.merge_cells("A1:E1")

    h2 = ["Objekt", "max. Beleihung %", "max. Beleihung EUR", "Substanzabschlag bei Eilverkauf %", "Verlust bei Eilverkauf EUR"]
    for i, h in enumerate(h2, start=1):
        style_header(ws2.cell(row=3, column=i, value=h))

    r2 = [
        ("Mehrfamilienhaus Bamberg", 60, 1677000, 10, 279500),
        ("Gewerbeimmobilie Forchheim", 60, 906000, 17, 256700),
        ("ETW Muenchen", 60, 567000, 6, 56700),
    ]
    for i, r in enumerate(r2, start=4):
        for j, v in enumerate(r, start=1):
            c = ws2.cell(row=i, column=j, value=v)
            style_cell(c, bold=(j == 1))
            if j in (3, 5):
                c.number_format = "#,##0 [$EUR]"
            if j in (2, 4):
                c.number_format = "0\"%\""

    auto_width(ws2, [32, 22, 22, 30, 26])

    wb.save(OUT / "liquiditaetsbetrachtung_immobilien_schaefer.xlsx")


if __name__ == "__main__":
    build_versorgungsausgleich()
    build_liquiditaetsbetrachtung()
    print("OK")
